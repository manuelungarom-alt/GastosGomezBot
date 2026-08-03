import os
import re
import json
import logging
import unicodedata
import difflib
from io import BytesIO
from datetime import datetime

import openpyxl
import dropbox
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("bot_familia")

# ---------------------------------------------------------------------------
# Config (todo esto viene de variables de entorno, nunca hardcodeado)
# ---------------------------------------------------------------------------
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
DROPBOX_REFRESH_TOKEN = os.environ["DROPBOX_REFRESH_TOKEN"]
DROPBOX_APP_KEY = os.environ["DROPBOX_APP_KEY"]
DROPBOX_APP_SECRET = os.environ["DROPBOX_APP_SECRET"]
EXCEL_PATH = os.environ.get("EXCEL_PATH", "/Gastos_Familia_v3.xlsx")
USERS_PATH = os.environ.get("USERS_PATH", "/usuarios.json")

ADMIN_NAMES = ["Luis", "Vero"]
PUEDE_CREDITO = ["Luis", "Vero", "Malena"]
ALL_NAMES = ["Luis", "Vero", "Malena", "Luka", "Mateo"]

FONDOS = {
    "vacaciones": "Vacaciones",
    "jubilacion": "Jubilación",
    "jubilación": "Jubilación",
    "emergencia": "Fondo de Emergencia",
}

MEDIOS = {
    "efectivo": "Efectivo",
    "mercado pago": "Mercado Pago",
    "mercadopago": "Mercado Pago",
    "mp": "Mercado Pago",
    "galicia": "Galicia",
    "santander": "Santander",
}
LUGARES = ["Efectivo", "Mercado Pago", "Galicia", "Santander"]
LUGARES_TRANSFERENCIA = ["Mercado Pago", "Galicia", "Santander"]

RETIRO_WORDS = ["retiro", "retiré", "retire", "saque", "saqué", "saco"]
INGRESO_WORDS = ["ingreso", "ingresé", "ingrese", "cobre", "cobré", "cobro"]
CREDITO_WORDS = ["credito", "crédito"]
PALABRAS_CLAVE = RETIRO_WORDS + INGRESO_WORDS + CREDITO_WORDS + list(FONDOS.keys())


def detectar_palabra_sospechosa(texto_sa: str) -> str | None:
    """Si hay una palabra parecida a una clave (ingreso/retiro/etc) pero mal escrita, la devuelve."""
    for palabra in re.findall(r"[a-záéíóúñ]+", texto_sa):
        if palabra in PALABRAS_CLAVE or len(palabra) < 4:
            continue
        match = difflib.get_close_matches(palabra, PALABRAS_CLAVE, n=1, cutoff=0.8)
        if match:
            return match[0]
    return None

TARJETAS = ["Visa", "American"]
BANCOS = ["Galicia", "Santander"]

# Estados de la conversacion de credito
TARJETA, BANCO, CUOTAS, MES, MEDIO_SELECT, TRANSFER_ORIGEN, MOTIVO_INPUT, TIPO_SELECT, LUGAR_INGRESO_SELECT = range(9)

dbx = dropbox.Dropbox(
    oauth2_refresh_token=DROPBOX_REFRESH_TOKEN,
    app_key=DROPBOX_APP_KEY,
    app_secret=DROPBOX_APP_SECRET,
)

# ---------------------------------------------------------------------------
# Dropbox helpers
# ---------------------------------------------------------------------------
def load_users() -> dict:
    try:
        _, res = dbx.files_download(USERS_PATH)
        return json.loads(res.content.decode("utf-8"))
    except Exception as e:
        log.warning(f"No se pudo leer {USERS_PATH} (puede ser normal si aun no existe): {e}")
        return {}


def save_users(users: dict):
    dbx.files_upload(
        json.dumps(users, ensure_ascii=False, indent=2).encode("utf-8"),
        USERS_PATH,
        mode=dropbox.files.WriteMode.overwrite,
    )


def download_excel() -> openpyxl.Workbook:
    _, res = dbx.files_download(EXCEL_PATH)
    return openpyxl.load_workbook(BytesIO(res.content))


def upload_excel(wb: openpyxl.Workbook):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    dbx.files_upload(buf.read(), EXCEL_PATH, mode=dropbox.files.WriteMode.overwrite)


def first_empty_row(ws, start=3, col=1) -> int:
    r = start
    while ws.cell(row=r, column=col).value not in (None, ""):
        r += 1
    return r


# ---------------------------------------------------------------------------
# Parseo del mensaje
# ---------------------------------------------------------------------------
ACCENT_MAP = {"a": "[aá]", "e": "[eé]", "i": "[ií]", "o": "[oó]", "u": "[uúü]"}


def patron_sin_acentos(palabra: str) -> str:
    return "".join(ACCENT_MAP.get(c, re.escape(c)) for c in palabra)


def quitar_acentos(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def parse_monto(s: str) -> float:
    """Interpreta '50.000' (miles) o '2479054,80' (decimales) correctamente."""
    partes = re.split(r"[.,]", s)
    if len(partes) == 1:
        return float(partes[0])
    ultima = partes[-1]
    if len(ultima) in (1, 2):
        entero = "".join(partes[:-1])
        return float(f"{entero}.{ultima}")
    return float("".join(partes))


def parse_message(text: str) -> dict | None:
    text_low = text.lower()
    text_low_sa = quitar_acentos(text_low)

    m = re.search(r"\$?\s*(\d+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)", text)
    if not m:
        return None
    try:
        monto = parse_monto(m.group(1))
    except ValueError:
        return None

    fondo = None
    for k, v in FONDOS.items():
        if k in text_low_sa:
            fondo = v
            break

    is_retiro = any(w in text_low_sa for w in RETIRO_WORDS)
    is_ingreso = any(w in text_low_sa for w in INGRESO_WORDS)
    is_credito = any(w in text_low_sa for w in CREDITO_WORDS)

    palabra_sospechosa = None
    if not (is_retiro or is_ingreso or is_credito or fondo):
        palabra_sospechosa = detectar_palabra_sospechosa(text_low_sa)

    medio = "Sin especificar"  # si no dicen el lugar, no se cuenta en la diversificación
    medio_encontrado = None
    for k, v in MEDIOS.items():
        if k in text_low:
            medio = v
            medio_encontrado = k
            break

    # Motivo: el mensaje original, sacando el monto y las palabras clave usadas
    motivo = text.replace(m.group(0), "")
    palabras_a_sacar = list(FONDOS.keys()) + RETIRO_WORDS + INGRESO_WORDS + CREDITO_WORDS
    if medio_encontrado:
        palabras_a_sacar.append(medio_encontrado)
    for palabra in palabras_a_sacar:
        motivo = re.sub(rf"\b{patron_sin_acentos(palabra)}\b", "", motivo, flags=re.IGNORECASE)
    motivo = re.sub(r"\s+", " ", motivo).strip(" ,.-")
    if not motivo:
        motivo = "Sin descripción"

    return {
        "monto": monto,
        "fondo": fondo,
        "is_retiro": is_retiro,
        "is_ingreso": is_ingreso,
        "is_credito": is_credito,
        "palabra_sospechosa": palabra_sospechosa,
        "medio": medio,
        "medio_detectado": medio_encontrado is not None,
        "motivo": motivo,
    }


# ---------------------------------------------------------------------------
# Escritura en Excel
# ---------------------------------------------------------------------------
def registrar_movimiento(nombre: str, tipo: str, monto: float, motivo: str, medio: str, fecha_override: datetime = None):
    wb = download_excel()
    ws = wb["Movimientos"]
    r = first_empty_row(ws, start=3, col=1)
    fecha = fecha_override or datetime.now()
    ws.cell(row=r, column=1, value=fecha)
    ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
    ws.cell(row=r, column=2, value=tipo)
    ws.cell(row=r, column=3, value=monto)
    ws.cell(row=r, column=3).number_format = "$#,##0"
    ws.cell(row=r, column=4, value=motivo)
    ws.cell(row=r, column=5, value=medio)
    ws.cell(row=r, column=6, value=nombre)
    ws.cell(row=r, column=7, value=f'=IF(A{r}="","",TEXT(A{r},"YYYY-MM"))')
    upload_excel(wb)


def registrar_ahorro(nombre: str, fondo: str, tipo: str, monto: float):
    wb = download_excel()
    ws = wb["Ahorro_Detalle"]
    r = first_empty_row(ws, start=3, col=1)
    ws.cell(row=r, column=1, value=datetime.now())
    ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
    ws.cell(row=r, column=2, value=fondo)
    ws.cell(row=r, column=3, value=tipo)
    ws.cell(row=r, column=4, value=monto)
    ws.cell(row=r, column=4).number_format = "$#,##0"
    ws.cell(row=r, column=5, value=nombre)
    ws.cell(row=r, column=6, value=f'=IF(A{r}="","",TEXT(A{r},"YYYY-MM"))')
    upload_excel(wb)


MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def parse_mes_input(texto: str) -> str | None:
    """Convierte 'agosto' o 'agosto2026' en 'YYYY-MM'."""
    texto_sa = quitar_acentos(texto.strip().lower())
    m = re.match(r"([a-z]+)\s*(\d{4})?", texto_sa)
    if not m:
        return None
    nombre_mes, anio_str = m.group(1), m.group(2)
    mes_num = MESES_ES.get(nombre_mes)
    if not mes_num:
        return None
    if anio_str:
        anio = int(anio_str)
    else:
        hoy = datetime.now()
        anio = hoy.year if mes_num >= hoy.month else hoy.year + 1
    return f"{anio:04d}-{mes_num:02d}"


def calcular_totales_mes(mes_sheet: str) -> dict:
    wb = download_excel()
    ws = wb["Movimientos"]
    ingresos = egresos = 0.0
    r = 3
    while ws.cell(row=r, column=1).value not in (None, ""):
        fecha = ws.cell(row=r, column=1).value
        if fecha and fecha.strftime("%Y-%m") == mes_sheet:
            tipo = ws.cell(row=r, column=2).value
            monto = ws.cell(row=r, column=3).value or 0
            if tipo == "Ingreso":
                ingresos += monto
            elif tipo == "Egreso":
                egresos += monto
        r += 1
    return {"ingresos": ingresos, "egresos": egresos, "balance": ingresos - egresos}


def mes_sheet_name(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def calcular_gasto_persona_mes(nombre_persona: str, mes_sheet: str) -> float:
    wb = download_excel()
    ws = wb["Movimientos"]
    total = 0.0
    r = 3
    while ws.cell(row=r, column=1).value not in (None, ""):
        fecha = ws.cell(row=r, column=1).value
        tipo = ws.cell(row=r, column=2).value
        persona = ws.cell(row=r, column=6).value
        if (
            fecha and fecha.strftime("%Y-%m") == mes_sheet
            and tipo == "Egreso"
            and persona == nombre_persona
        ):
            total += ws.cell(row=r, column=3).value or 0
        r += 1
    return total


def calcular_totales() -> dict:
    """Recalcula a mano (no depende de la cache de formulas de Excel)."""
    wb = download_excel()

    ws = wb["Movimientos"]
    ingresos = egresos = 0.0
    lugares_total = {l: 0.0 for l in LUGARES}
    sin_especificar = 0.0
    r = 3
    while ws.cell(row=r, column=1).value not in (None, ""):
        tipo = ws.cell(row=r, column=2).value
        monto = ws.cell(row=r, column=3).value or 0
        medio = ws.cell(row=r, column=5).value
        if tipo == "Ingreso":
            ingresos += monto
            if medio in lugares_total:
                lugares_total[medio] += monto
            else:
                sin_especificar += monto
        elif tipo == "Egreso":
            egresos += monto
            if medio in lugares_total:
                lugares_total[medio] -= monto
            else:
                sin_especificar -= monto
        r += 1

    wsd = wb["Ahorro_Detalle"]
    fondos_total = {"Vacaciones": 0.0, "Jubilación": 0.0, "Fondo de Emergencia": 0.0}
    r = 3
    while wsd.cell(row=r, column=1).value not in (None, ""):
        fondo = wsd.cell(row=r, column=2).value
        tipo = wsd.cell(row=r, column=3).value
        monto = wsd.cell(row=r, column=4).value or 0
        if fondo in fondos_total:
            fondos_total[fondo] += monto if tipo == "Aporte" else -monto
        r += 1

    return {
        "ingresos": ingresos,
        "egresos": egresos,
        "balance": ingresos - egresos,
        "fondos": fondos_total,
        "lugares": lugares_total,
        "sin_especificar": sin_especificar,
    }


def fmt(n: float) -> str:
    return f"${n:,.0f}".replace(",", ".")


# ---------------------------------------------------------------------------
# Handlers de Telegram
# ---------------------------------------------------------------------------
async def reset_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    users = load_users()
    if chat_id in users:
        del users[chat_id]
        save_users(users)
    await start(update, context)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    users = load_users()
    if chat_id in users:
        await update.message.reply_text(f"Hola {users[chat_id]}, ya te tengo registrado ✅")
        return
    buttons = [[InlineKeyboardButton(n, callback_data=f"reg:{n}")] for n in ALL_NAMES]
    await update.message.reply_text(
        "¡Hola! ¿Quién sos?", reply_markup=InlineKeyboardMarkup(buttons)
    )


async def register_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    name = query.data.split(":")[1]
    chat_id = str(query.message.chat_id)
    users = load_users()
    users[chat_id] = name
    save_users(users)
    await query.answer()
    await query.edit_message_text(f"Listo, quedaste registrado como {name} ✅")


async def total_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    users = load_users()
    nombre = users.get(chat_id)
    if nombre not in ADMIN_NAMES:
        await update.message.reply_text("Este comando es solo para Luis y Vero.")
        return

    texto_comando = update.message.text.split()[0][1:]  # sin la barra "/"
    sufijo = texto_comando[len("total"):].strip()

    if sufijo:
        mes_sheet = parse_mes_input(sufijo)
        if mes_sheet:
            tm = calcular_totales_mes(mes_sheet)
            msg = (
                f"📊 *{mes_sheet}*\n"
                f"Ingresos: {fmt(tm['ingresos'])}\n"
                f"Egresos: {fmt(tm['egresos'])}\n"
                f"Balance del mes: {fmt(tm['balance'])}\n"
            )
            await update.message.reply_text(msg, parse_mode="Markdown")
            return

        sufijo_norm = quitar_acentos(sufijo.lower())
        persona_encontrada = None
        for persona in ALL_NAMES:
            if quitar_acentos(persona.lower()) == sufijo_norm:
                persona_encontrada = persona
                break
        if persona_encontrada:
            mes_actual = mes_sheet_name(datetime.now())
            gasto = calcular_gasto_persona_mes(persona_encontrada, mes_actual)
            await update.message.reply_text(
                f"{persona_encontrada} gastó {fmt(gasto)} en {mes_actual}"
            )
            return

        await update.message.reply_text(
            f"No entendí '{sufijo}'. Probá '/totalagosto' (un mes) o '/total{sufijo.lower()}' con el nombre de alguien de la familia."
        )
        return

    t = calcular_totales()
    ahorrado_total = sum(t["fondos"].values())
    balance_disponible = t["balance"] - ahorrado_total
    msg = (
        f"📊 *Resumen*\n"
        f"Ingresos: {fmt(t['ingresos'])}\n"
        f"Egresos: {fmt(t['egresos'])}\n"
        f"Ahorrado (separado): {fmt(ahorrado_total)}\n"
        f"Balance disponible: {fmt(balance_disponible)}\n\n"
        f"📍 *Dónde está la plata*\n"
        f"Efectivo: {fmt(t['lugares']['Efectivo'])}\n"
        f"Mercado Pago: {fmt(t['lugares']['Mercado Pago'])}\n"
        f"Galicia: {fmt(t['lugares']['Galicia'])}\n"
        f"Santander: {fmt(t['lugares']['Santander'])}\n"
        f"Sin especificar: {fmt(t['sin_especificar'])}\n\n"
        f"💰 *Ahorros*\n"
        f"Fondo de Emergencia: {fmt(t['fondos']['Fondo de Emergencia'])}\n"
        f"Jubilación: {fmt(t['fondos']['Jubilación'])}\n"
        f"Vacaciones: {fmt(t['fondos']['Vacaciones'])}\n"
    )
    await update.message.reply_text(msg, parse_mode="Markdown")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    users = load_users()
    nombre = users.get(chat_id)

    if not nombre:
        await update.message.reply_text("Antes que nada, decime quién sos con /start")
        return ConversationHandler.END

    text = update.message.text or ""
    parsed = parse_message(text)
    if not parsed:
        await update.message.reply_text(
            "No reconocí ese mensaje como un gasto/ingreso. Probá algo como: '50.000 comida transferencia'"
        )
        return ConversationHandler.END

    if parsed["palabra_sospechosa"]:
        await update.message.reply_text(
            f"⚠️ No entendí bien tu mensaje — ¿quisiste decir '{parsed['palabra_sospechosa']}'? "
            f"No cargué nada, revisá cómo lo escribiste y mandalo de nuevo."
        )
        return ConversationHandler.END

    es_admin = nombre in ADMIN_NAMES

    # --- Credito (Luis, Vero, Malena; solo gastos) ---
    if parsed["is_credito"] and not parsed["is_ingreso"] and not parsed["fondo"] and nombre in PUEDE_CREDITO:
        context.user_data["pendiente"] = {
            "nombre": nombre,
            "monto": parsed["monto"],
            "motivo": parsed["motivo"],
        }
        buttons = [[InlineKeyboardButton(t, callback_data=f"tarjeta:{t}")] for t in TARJETAS]
        await update.message.reply_text(
            "¿Qué tarjeta usaste?", reply_markup=InlineKeyboardMarkup(buttons)
        )
        return TARJETA

    # --- Ahorro (solo admins) ---
    if parsed["fondo"] and es_admin:
        tipo = "Retiro" if parsed["is_retiro"] else "Aporte"
        registrar_ahorro(nombre, parsed["fondo"], tipo, parsed["monto"])
        await update.message.reply_text(
            f"Listo ✅ {tipo} de {fmt(parsed['monto'])} en {parsed['fondo']}"
        )
        return ConversationHandler.END

    # --- Mensaje ambiguo: solo el monto, nada mas. Preguntar tipo antes de asumir ---
    if (
        parsed["motivo"] == "Sin descripción"
        and not parsed["medio_detectado"]
        and not parsed["is_ingreso"]
        and es_admin
    ):
        context.user_data["pendiente"] = {"nombre": nombre, "monto": parsed["monto"], "es_admin": es_admin}
        buttons = [[
            InlineKeyboardButton("Ingreso", callback_data="tiposel:Ingreso"),
            InlineKeyboardButton("Egreso", callback_data="tiposel:Egreso"),
        ]]
        await update.message.reply_text(
            "¿Ingreso o egreso?", reply_markup=InlineKeyboardMarkup(buttons)
        )
        return TIPO_SELECT

    # --- Ingreso (solo admins) ---
    if parsed["is_ingreso"] and es_admin:
        context.user_data["pendiente"] = {
            "nombre": nombre, "monto": parsed["monto"], "motivo": parsed["motivo"],
            "tipo": "Ingreso", "medio": parsed["medio"], "medio_detectado": parsed["medio_detectado"],
        }
        if parsed["motivo"] == "Sin descripción":
            return await preguntar_motivo(update, context)
        if not parsed["medio_detectado"]:
            return await preguntar_medio(update, context)
        return await finalizar_ingreso_egreso(update, context)

    # --- Egreso normal (todos) ---
    context.user_data["pendiente"] = {
        "nombre": nombre, "monto": parsed["monto"], "motivo": parsed["motivo"],
        "tipo": "Egreso", "medio": parsed["medio"], "medio_detectado": parsed["medio_detectado"],
        "es_admin": es_admin,
    }
    if parsed["motivo"] == "Sin descripción":
        return await preguntar_motivo(update, context)
    if not parsed["medio_detectado"]:
        return await preguntar_medio(update, context)
    return await finalizar_ingreso_egreso(update, context)


# ---------------------------------------------------------------------------
# Pregunta de medio de pago: Transferencia -> (de donde) / Efectivo / Credito
# ---------------------------------------------------------------------------
async def tipo_select_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    tipo = query.data.split(":", 1)[1]
    await query.answer()
    pend = context.user_data["pendiente"]
    pend["tipo"] = tipo

    if tipo == "Ingreso":
        await query.edit_message_text("Tipo: Ingreso")
        buttons = [[InlineKeyboardButton(l, callback_data=f"lugaringreso:{l}")] for l in LUGARES]
        await query.message.reply_text(
            "¿En qué lugar ingresó?", reply_markup=InlineKeyboardMarkup(buttons)
        )
        return LUGAR_INGRESO_SELECT

    pend["motivo"] = "Sin descripción"
    pend["medio_detectado"] = False
    await query.edit_message_text("Tipo: Egreso")
    return await preguntar_motivo(update, context)


async def lugar_ingreso_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    lugar = query.data.split(":", 1)[1]
    await query.answer()
    pend = context.user_data["pendiente"]
    pend["medio"] = lugar
    pend["motivo"] = "Sin descripción"
    await query.edit_message_text(f"Ingresó en: {lugar}")
    return await finalizar_ingreso_egreso(update, context)


async def preguntar_motivo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.effective_message.reply_text("¿En qué fue?")
    return MOTIVO_INPUT


async def motivo_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
    if not texto:
        await update.message.reply_text("Contame en qué fue, por favor.")
        return MOTIVO_INPUT
    pend = context.user_data["pendiente"]
    pend["motivo"] = texto
    if not pend["medio_detectado"]:
        return await preguntar_medio(update, context)
    return await finalizar_ingreso_egreso(update, context)


async def preguntar_medio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    pend = context.user_data["pendiente"]
    if pend["nombre"] in PUEDE_CREDITO:
        opciones = ["Transferencia", "Efectivo"]
        if pend["tipo"] == "Egreso":
            opciones.append("Crédito")
    else:
        # Mateo y Luka: solo estas 2 opciones, sin submenu de banco
        opciones = ["Efectivo", "Mercado Pago"]
    buttons = [[InlineKeyboardButton(o, callback_data=f"mediosel:{o}")] for o in opciones]
    await update.effective_message.reply_text(
        "¿Medio de pago?", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return MEDIO_SELECT


async def medio_select_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    opcion = query.data.split(":", 1)[1]
    await query.answer()
    pend = context.user_data["pendiente"]

    if opcion == "Transferencia":
        await query.edit_message_text("Transferencia\n¿De dónde?")
        buttons = [[InlineKeyboardButton(b, callback_data=f"transorigen:{b}")] for b in LUGARES_TRANSFERENCIA]
        await query.message.reply_text("Elegí de dónde:", reply_markup=InlineKeyboardMarkup(buttons))
        return TRANSFER_ORIGEN

    if opcion == "Crédito":
        await query.edit_message_text("Medio: Crédito")
        buttons = [[InlineKeyboardButton(t, callback_data=f"tarjeta:{t}")] for t in TARJETAS]
        await query.message.reply_text("¿Qué tarjeta usaste?", reply_markup=InlineKeyboardMarkup(buttons))
        return TARJETA

    # Efectivo o Mercado Pago (elegido directo, sin pasar por el submenu de banco)
    pend["medio"] = opcion
    await query.edit_message_text(f"Medio: {opcion}")
    return await finalizar_ingreso_egreso(update, context)


async def transfer_origen_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    lugar = query.data.split(":", 1)[1]
    await query.answer()
    context.user_data["pendiente"]["medio"] = lugar
    await query.edit_message_text(f"Transferencia desde: {lugar}")
    return await finalizar_ingreso_egreso(update, context)


async def finalizar_ingreso_egreso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    pend = context.user_data["pendiente"]
    registrar_movimiento(pend["nombre"], pend["tipo"], pend["monto"], pend["motivo"], pend["medio"])
    if pend["tipo"] == "Ingreso":
        await update.effective_message.reply_text(f"Listo ✅ Ingreso de {fmt(pend['monto'])} anotado")
    elif pend.get("es_admin", True):
        await update.effective_message.reply_text(
            f"Listo ✅ Gasto de {fmt(pend['monto'])} en {pend['motivo']} ({pend['medio']})"
        )
    else:
        await update.effective_message.reply_text("Listo, anotado ✅")
    context.user_data.pop("pendiente", None)
    return ConversationHandler.END


# ---------------------------------------------------------------------------
# Conversacion de credito: tarjeta -> banco -> cuotas -> mes
# ---------------------------------------------------------------------------
def sumar_meses(mes_sheet: str, n: int) -> str:
    anio, mes = map(int, mes_sheet.split("-"))
    total = (anio * 12 + (mes - 1)) + n
    return f"{total // 12:04d}-{total % 12 + 1:02d}"


async def tarjeta_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    tarjeta = query.data.split(":")[1]
    context.user_data["pendiente"]["tarjeta"] = tarjeta
    await query.answer()
    buttons = [[InlineKeyboardButton(b, callback_data=f"banco:{b}")] for b in BANCOS]
    await query.edit_message_text(f"Tarjeta: {tarjeta}\n¿De qué banco?")
    await query.message.reply_text("Elegí el banco:", reply_markup=InlineKeyboardMarkup(buttons))
    return BANCO


async def banco_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    banco = query.data.split(":")[1]
    context.user_data["pendiente"]["banco"] = banco
    await query.answer()
    await query.edit_message_text(f"Banco: {banco}\n¿Cuántas cuotas? (escribí el número)")
    return CUOTAS


async def cuotas_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
    if not texto.isdigit() or int(texto) < 1:
        await update.message.reply_text("Escribí un número válido de cuotas (ej. 1, 3, 12).")
        return CUOTAS
    context.user_data["pendiente"]["cuotas"] = int(texto)
    await update.message.reply_text(
        "¿En qué mes empieza a cobrarse la primera cuota? (ej. 'agosto' o 'agosto 2026')"
    )
    return MES


async def mes_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
    mes_sheet = parse_mes_input(texto)
    if not mes_sheet:
        await update.message.reply_text("No entendí el mes. Escribilo así: 'agosto' o 'agosto 2026'.")
        return MES

    pend = context.user_data.get("pendiente")
    if not pend:
        await update.message.reply_text("Se perdió el gasto en curso, probá de nuevo desde cero.")
        return ConversationHandler.END

    cuotas = pend["cuotas"]
    monto_total = pend["monto"]
    monto_cuota = round(monto_total / cuotas)
    diferencia = monto_total - monto_cuota * cuotas

    for i in range(cuotas):
        monto_i = monto_cuota + (diferencia if i == cuotas - 1 else 0)
        mes_i = sumar_meses(mes_sheet, i)
        anio_i, mes_num_i = map(int, mes_i.split("-"))
        fecha_i = datetime(anio_i, mes_num_i, 1)
        motivo_i = f"{pend['motivo']} ({pend['tarjeta']}, cuota {i + 1}/{cuotas})"
        registrar_movimiento(pend["nombre"], "Egreso", monto_i, motivo_i, pend["banco"], fecha_override=fecha_i)

    await update.message.reply_text(
        f"Listo ✅ {fmt(monto_total)} en {cuotas} cuota(s) de {fmt(monto_cuota)}, "
        f"empezando en {mes_sheet} ({pend['tarjeta']} {pend['banco']})"
    )
    context.user_data.pop("pendiente", None)
    return ConversationHandler.END


async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("pendiente", None)
    await update.message.reply_text("Cancelado.")
    return ConversationHandler.END


# ---------------------------------------------------------------------------
# Arranque en modo webhook (Telegram nos avisa a nosotros, en vez de que
# nosotros le preguntemos todo el tiempo). Esto es clave para el plan free
# de Render: un pedido HTTP entrante es lo unico que puede "despertar" al
# servicio si estaba dormido por inactividad.
# ---------------------------------------------------------------------------
def main():
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message)],
        states={
            TIPO_SELECT: [CallbackQueryHandler(tipo_select_callback, pattern=r"^tiposel:")],
            LUGAR_INGRESO_SELECT: [CallbackQueryHandler(lugar_ingreso_callback, pattern=r"^lugaringreso:")],
            MOTIVO_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, motivo_input_handler)],
            MEDIO_SELECT: [CallbackQueryHandler(medio_select_callback, pattern=r"^mediosel:")],
            TRANSFER_ORIGEN: [CallbackQueryHandler(transfer_origen_callback, pattern=r"^transorigen:")],
            TARJETA: [CallbackQueryHandler(tarjeta_callback, pattern=r"^tarjeta:")],
            BANCO: [CallbackQueryHandler(banco_callback, pattern=r"^banco:")],
            CUOTAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, cuotas_handler)],
            MES: [MessageHandler(filters.TEXT & ~filters.COMMAND, mes_handler)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("reset", reset_cmd))
    application.add_handler(MessageHandler(filters.Regex(r"^/total") & filters.COMMAND, total_cmd))
    application.add_handler(CallbackQueryHandler(register_callback, pattern=r"^reg:"))
    application.add_handler(conv_handler)

    port = int(os.environ.get("PORT", 10000))
    webhook_base = os.environ["WEBHOOK_URL"].rstrip("/")

    log.info("Bot arrancado en modo webhook...")
    application.run_webhook(
        listen="0.0.0.0",
        port=port,
        url_path=TELEGRAM_TOKEN,
        webhook_url=f"{webhook_base}/{TELEGRAM_TOKEN}",
    )


if __name__ == "__main__":
    main()
